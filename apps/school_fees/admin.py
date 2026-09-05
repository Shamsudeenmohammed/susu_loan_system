import csv
import io

from django.contrib import admin, messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .models import (
    SchoolClass,
    AcademicYear,
    Term,
    Student,
    FeeCategory,
    FeeStructure,
    StudentFeeAccount,
    FeePayment,
    ReminderTemplate,
    ReminderLog,
)


@admin.register(SchoolClass)
class SchoolClassAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'is_active', 'created_at']
    list_filter = ['is_active']
    search_fields = ['name', 'code']


@admin.register(AcademicYear)
class AcademicYearAdmin(admin.ModelAdmin):
    list_display = ['name', 'start_date', 'end_date', 'is_active', 'created_at']
    list_filter = ['is_active']
    search_fields = ['name']


@admin.register(Term)
class TermAdmin(admin.ModelAdmin):
    list_display = ['academic_year', 'name', 'term_number', 'start_date', 'end_date']
    list_filter = ['academic_year', 'term_number']
    search_fields = ['name', 'academic_year__name']


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = ['student_id', 'first_name', 'last_name', 'school_class', 'is_active', 'created_at']
    list_filter = ['school_class', 'is_active']
    search_fields = ['student_id', 'first_name', 'last_name', 'parent_name', 'parent_phone']
    readonly_fields = ['student_id', 'created_at', 'updated_at']
    autocomplete_fields = ['school_class']

    REQUIRED_FIELDS = ['first_name', 'last_name', 'school_class', 'parent_name', 'parent_phone']
    OPTIONAL_FIELDS = ['parent_email', 'is_active']
    HEADER_ALIASES = {
        'firstname': 'first_name',
        'first_name': 'first_name',
        'first name': 'first_name',
        'lastname': 'last_name',
        'last_name': 'last_name',
        'last name': 'last_name',
        'schoolclass': 'school_class',
        'school_class': 'school_class',
        'school class': 'school_class',
        'class': 'school_class',
        'classname': 'school_class',
        'class name': 'school_class',
        'parentname': 'parent_name',
        'parent_name': 'parent_name',
        'parent name': 'parent_name',
        'parentphone': 'parent_phone',
        'parent_phone': 'parent_phone',
        'parent phone': 'parent_phone',
        'phone': 'parent_phone',
        'phonenumber': 'parent_phone',
        'parentemail': 'parent_email',
        'parent_email': 'parent_email',
        'parent email': 'parent_email',
        'email': 'parent_email',
        'isactive': 'is_active',
        'is_active': 'is_active',
        'active': 'is_active',
    }

    actions = ['import_students']

    @admin.action(description='Import students from file (CSV/Excel)')
    def import_students(self, request, queryset):
        return HttpResponseRedirect(reverse('admin:school_fees_student_import'))

    import_students.allows_empty = True

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-students/',
                self.admin_site.admin_view(self.import_students_view),
                name='school_fees_student_import',
            ),
            path(
                'import-students/template.csv',
                self.admin_site.admin_view(self.download_import_template),
                name='school_fees_student_import_template',
            ),
        ]
        return custom_urls + urls

    def has_import_permission(self, request):
        return self.has_add_permission(request)

    def import_students_view(self, request):
        if not self.has_import_permission(request):
            raise PermissionDenied
        results = None
        if request.method == 'POST':
            upload = request.FILES.get('file')
            if not upload:
                messages.error(request, 'Please select a CSV or Excel (.xlsx) file to upload.')
            else:
                results = self._process_upload(request, upload)
        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'title': 'Import students',
            'results': results,
        }
        return render(request, 'admin/school_fees/student/import_students.html', context)

    def download_import_template(self, request):
        if not self.has_import_permission(request):
            raise PermissionDenied
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="student_import_template.csv"'
        writer = csv.writer(response)
        writer.writerow(self.REQUIRED_FIELDS + self.OPTIONAL_FIELDS)
        writer.writerow(['Akuapem', 'Kufuor', 'JHS 2', 'Mr. Kufuor', '0244000001', '', '1'])
        writer.writerow(['Adjoa', 'Mensah', 'Primary 4', 'Mrs. Mensah', '0203000002', 'parent@example.com', '1'])
        return response

    def _parse_csv(self, upload):
        try:
            text = upload.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            text = upload.read().decode('latin-1')
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], 'The CSV file has no header row.'
        rows = []
        for row in reader:
            if row is None or all(not (v or '').strip() for v in row.values()):
                continue
            mapped = {}
            for key, value in row.items():
                canonical = self.HEADER_ALIASES.get((key or '').strip().lower())
                if canonical:
                    mapped[canonical] = value.strip() if isinstance(value, str) else value
            if mapped:
                rows.append(mapped)
        return rows, None

    def _parse_xlsx(self, upload):
        if not HAS_OPENPYXL:
            return [], 'openpyxl is not installed. Cannot parse XLSX files.'
        wb = None
        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            first = True
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() if v is not None else '' for v in row]
                if first:
                    headers = [v.lower() for v in values]
                    first = False
                    if not headers:
                        return [], 'The XLSX file has no header row.'
                    continue
                if all(not v for v in values):
                    continue
                mapped = {}
                for idx, value in enumerate(values):
                    if idx < len(headers):
                        canonical = self.HEADER_ALIASES.get(headers[idx])
                        if canonical:
                            mapped[canonical] = value
                if mapped:
                    rows.append(mapped)
            return rows, None
        except Exception:
            return [], 'The XLSX file could not be read. Make sure it is a valid .xlsx file.'
        finally:
            if wb is not None:
                wb.close()

    def _parse_bool(self, value, default):
        if value is None or value == '':
            return default
        v = str(value).strip().lower()
        if v in ('1', 'true', 'yes', 'y', 'on', 'active'):
            return True
        if v in ('0', 'false', 'no', 'n', 'off', 'inactive'):
            return False
        return default

    def _process_upload(self, request, upload):
        name = upload.name.lower()
        if name.endswith('.csv'):
            rows, error = self._parse_csv(upload)
        elif name.endswith('.xlsx'):
            rows, error = self._parse_xlsx(upload)
        else:
            return {'fatal': 'Unsupported file type. Please upload a .csv or .xlsx file.'}
        if error:
            return {'fatal': error}
        if not rows:
            return {'fatal': 'The file contains no data rows.'}

        created, skipped, problems = [], [], []
        seen = set()
        for idx, row in enumerate(rows, start=2):
            missing = [f for f in self.REQUIRED_FIELDS if not (row.get(f) or '').strip()]
            if missing:
                problems.append((idx, f'Missing required field(s): {", ".join(missing)}'))
                continue
            class_ref = row['school_class'].strip()
            school_class = (
                SchoolClass.objects.filter(code__iexact=class_ref).first()
                or SchoolClass.objects.filter(name__iexact=class_ref).first()
            )
            if not school_class:
                problems.append((idx, f"SchoolClass '{class_ref}' not found (match by name or code)"))
                continue
            first_name = row['first_name'].strip()
            last_name = row['last_name'].strip()
            parent_email = row.get('parent_email') or ''
            dup_key = (first_name.lower(), last_name.lower(), school_class.pk)
            if dup_key in seen:
                problems.append((idx, 'Duplicate row within file (same first name, last name and class)'))
                continue
            exists = Student.objects.filter(
                first_name__iexact=first_name,
                last_name__iexact=last_name,
                school_class=school_class,
            ).exists()
            if exists:
                skipped.append(f"{first_name} {last_name} ({school_class.name})")
                continue
            try:
                Student.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    school_class=school_class,
                    parent_name=row['parent_name'].strip(),
                    parent_phone=row['parent_phone'].strip(),
                    parent_email=parent_email.strip(),
                    is_active=self._parse_bool(row.get('is_active'), True),
                    created_by=request.user,
                )
                created.append(f"{first_name} {last_name} ({school_class.name})")
                seen.add(dup_key)
            except Exception as exc:
                problems.append((idx, f'Could not create student: {exc}'))

        result = {
            'created_count': len(created),
            'skipped_count': len(skipped),
            'problems_count': len(problems),
            'created': created,
            'skipped': skipped,
            'problems': problems,
        }
        if not problems:
            messages.success(
                request,
                f"Imported {len(created)} student(s)"
                + (f" and skipped {len(skipped)} existing." if skipped else '.'),
            )
        return result


@admin.register(FeeCategory)
class FeeCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'is_active', 'created_at']
    list_filter = ['is_active']
    search_fields = ['name', 'code']


@admin.register(FeeStructure)
class FeeStructureAdmin(admin.ModelAdmin):
    list_display = ['academic_year', 'term', 'school_class', 'fee_category', 'amount', 'due_date', 'is_active']
    list_filter = ['academic_year', 'term', 'school_class', 'fee_category', 'is_active']
    search_fields = ['academic_year__name', 'school_class__name', 'fee_category__name', 'description']
    autocomplete_fields = ['academic_year', 'term', 'school_class', 'fee_category']


@admin.register(StudentFeeAccount)
class StudentFeeAccountAdmin(admin.ModelAdmin):
    list_display = [
        'account_number', 'student', 'academic_year', 'term', 'total_fees',
        'amount_paid', 'outstanding_balance', 'status', 'last_payment_date',
    ]
    list_filter = ['academic_year', 'term', 'status']
    search_fields = ['account_number', 'student__student_id', 'student__first_name', 'student__last_name']
    readonly_fields = ['account_number', 'outstanding_balance', 'created_at', 'updated_at']
    autocomplete_fields = ['student', 'academic_year', 'term']


@admin.register(FeePayment)
class FeePaymentAdmin(admin.ModelAdmin):
    list_display = [
        'receipt_number', 'student', 'account', 'amount', 'payment_date',
        'payment_method', 'reference', 'is_online', 'recorded_by', 'created_at',
    ]
    list_filter = ['payment_method', 'payment_date', 'academic_year', 'term', 'is_online']
    search_fields = ['receipt_number', 'reference', 'student__first_name', 'student__last_name']
    readonly_fields = ['receipt_number', 'previous_balance', 'remaining_balance', 'created_at']
    autocomplete_fields = ['student', 'account', 'academic_year', 'term', 'recorded_by']


@admin.register(ReminderTemplate)
class ReminderTemplateAdmin(admin.ModelAdmin):
    list_display = ['name', 'reminder_type', 'is_active', 'updated_at']
    list_filter = ['reminder_type', 'is_active']
    search_fields = ['name', 'message']


@admin.register(ReminderLog)
class ReminderLogAdmin(admin.ModelAdmin):
    list_display = ['student', 'parent_phone', 'reminder_type', 'status', 'created_at']
    list_filter = ['status', 'reminder_type', 'created_at']
    search_fields = ['student__first_name', 'student__last_name', 'parent_phone', 'message']
    readonly_fields = ['student', 'parent_phone', 'message', 'status', 'unique_key', 'created_at']
